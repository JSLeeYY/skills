from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.properties import CalcProperties


OUTPUT = Path(
    r"d:\DevelopmentLocation\agent skill\skills\word_project\2026\2026-04-20\软件项目供应商评审评分表.xlsx"
)
FALLBACK_OUTPUT = OUTPUT.with_name(f"{OUTPUT.stem}-单报价版{OUTPUT.suffix}")
DEDUP_OUTPUT = OUTPUT.with_name(f"{OUTPUT.stem}-去重版{OUTPUT.suffix}")
DEDUP_FALLBACK_OUTPUT = OUTPUT.with_name(f"{OUTPUT.stem}-去重版-v2{OUTPUT.suffix}")
CUSTOM_WEIGHT_DEFAULT = 60
OPS_WEIGHT_DEFAULT = 0
PRICE_WEIGHT_DEFAULT = 40
PRICE_SHEET_TITLE = "统一报价评分"


thin = Side(style="thin", color="7F7F7F")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

font_title = Font(name="宋体", size=16, bold=True)
font_subtitle = Font(name="宋体", size=11, bold=True)
font_header = Font(name="宋体", size=10, bold=True, color="FFFFFF")
font_body = Font(name="宋体", size=10)
font_input = Font(name="宋体", size=10, color="0000FF")
font_formula = Font(name="宋体", size=10, color="000000")

fill_title = PatternFill("solid", fgColor="D9EAF7")
fill_header = PatternFill("solid", fgColor="1F4E78")
fill_section = PatternFill("solid", fgColor="DDEBF7")
fill_note = PatternFill("solid", fgColor="FFF2CC")
fill_formula = PatternFill("solid", fgColor="F3F3F3")
fill_summary = PatternFill("solid", fgColor="E2F0D9")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_cell(cell, *, font=None, fill=None, alignment=None, border_on=True, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border_on:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def write_row(ws, row, values, *, font=None, fill=None, alignment=None, number_format=None):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        style_cell(
            cell,
            font=font or font_body,
            fill=fill,
            alignment=alignment or align_left,
            number_format=number_format,
        )


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def merged_title(ws, cell_range, value, *, fill=fill_title, font=font_title, height=24):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    style_cell(cell, font=font, fill=fill, alignment=align_center)
    ws.row_dimensions[cell.row].height = height


def is_user_value(value):
    return value is not None and not (isinstance(value, str) and value.startswith("="))


def load_existing_inputs(dev_count, ops_count):
    data = {
        "qualification": {},
        "dev_scores": {},
        "ops_scores": {},
        "supplier_headers": ["供应商A有效报价", "供应商B有效报价", "供应商C有效报价"],
        "quote_values": [None, None, None],
        "custom_weight": CUSTOM_WEIGHT_DEFAULT,
        "ops_weight": OPS_WEIGHT_DEFAULT,
        "quote_weight": PRICE_WEIGHT_DEFAULT,
    }
    source_path = None
    for candidate in (DEDUP_OUTPUT, FALLBACK_OUTPUT, OUTPUT):
        if candidate.exists():
            source_path = candidate
            break
    if source_path is None:
        return data

    try:
        wb = load_workbook(source_path, data_only=False)
    except Exception:
        return data

    if "资格审查" in wb.sheetnames:
        ws = wb["资格审查"]
        for row in range(5, 15):
            for col in range(4, 7):
                value = ws.cell(row, col).value
                if is_user_value(value):
                    data["qualification"][(row, col)] = value

    def capture_scores(sheet_name, count, target_key):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        for row in range(5, 5 + count):
            for col in range(6, 9):
                value = ws.cell(row, col).value
                if is_user_value(value):
                    data[target_key][(row, col)] = value
        for idx, ref in enumerate(("J2", "K2", "L2")):
            value = ws[ref].value
            if value:
                data["supplier_headers"][idx] = value
        for idx, ref in enumerate(("J3", "K3", "L3")):
            value = ws[ref].value
            if is_user_value(value):
                data["quote_values"][idx] = value

    capture_scores("定制开发评分", dev_count, "dev_scores")
    capture_scores("驻场运维评分", ops_count, "ops_scores")

    if PRICE_SHEET_TITLE in wb.sheetnames:
        ws = wb[PRICE_SHEET_TITLE]
        for idx, ref in enumerate(("J2", "K2", "L2")):
            value = ws[ref].value
            if value:
                data["supplier_headers"][idx] = value
        for idx, ref in enumerate(("J3", "K3", "L3")):
            value = ws[ref].value
            if is_user_value(value):
                data["quote_values"][idx] = value

    if "汇总排名" in wb.sheetnames:
        ws = wb["汇总排名"]
        title_a4 = ws["A4"].value
        if title_a4 == "定制开发权重（%）":
            if is_user_value(ws["B4"].value):
                data["custom_weight"] = ws["B4"].value
            if is_user_value(ws["B5"].value):
                data["ops_weight"] = ws["B5"].value
            if is_user_value(ws["B6"].value):
                data["quote_weight"] = ws["B6"].value
        elif title_a4 == "技术商务总权重（%）":
            tech_weight = ws["B4"].value if is_user_value(ws["B4"].value) else 60
            price_weight = None
            if is_user_value(ws["B5"].value):
                price_weight = ws["B5"].value
            elif isinstance(ws["B5"].value, str) and ws["B5"].value == "=100-B4":
                price_weight = round(100 - tech_weight, 2)
            dev_inner = ws["B6"].value if is_user_value(ws["B6"].value) else 100
            ops_inner = None
            if is_user_value(ws["B7"].value):
                ops_inner = ws["B7"].value
            elif isinstance(ws["B7"].value, str) and ws["B7"].value == "=100-B6":
                ops_inner = round(100 - dev_inner, 2)
            if price_weight is None:
                price_weight = 40
            if ops_inner is None:
                ops_inner = 0
            data["custom_weight"] = round(tech_weight * dev_inner / 100, 2)
            data["ops_weight"] = round(tech_weight * ops_inner / 100, 2)
            data["quote_weight"] = round(price_weight, 2)

    wb.close()
    return data


def apply_existing_inputs(wb, existing):
    ws = wb["资格审查"]
    for (row, col), value in existing["qualification"].items():
        ws.cell(row, col, value)

    for sheet_name, key in (("定制开发评分", "dev_scores"), ("驻场运维评分", "ops_scores")):
        ws = wb[sheet_name]
        input_last_row = ws.max_row - 1
        for (row, col), value in existing[key].items():
            if 5 <= row <= input_last_row:
                ws.cell(row, col, value)

    ws = wb[PRICE_SHEET_TITLE]
    for idx, ref in enumerate(("J2", "K2", "L2")):
        if existing["supplier_headers"][idx]:
            ws[ref] = existing["supplier_headers"][idx]
    for idx, ref in enumerate(("J3", "K3", "L3")):
        if existing["quote_values"][idx] is not None:
            ws[ref] = existing["quote_values"][idx]

    ws = wb["汇总排名"]
    ws["B4"] = existing["custom_weight"]
    ws["B5"] = existing["ops_weight"]
    ws["B6"] = existing["quote_weight"]


def build_integer_score_labels(max_points):
    max_points = int(max_points)
    if max_points <= 1:
        return [f"{max_points} 分", "0 分", "0 分"]
    mid_start = (max_points + 1) // 2
    mid_end = max_points - 1
    low_end = mid_start - 1
    mid_label = f"{mid_start} 分" if mid_start == mid_end else f"{mid_start}-{mid_end} 分"
    low_label = "0 分" if low_end <= 0 else (f"0-{low_end} 分" if low_end > 1 else "0-1 分")
    return [f"{max_points} 分", mid_label, low_label]


def rewrite_standard_scores(text, max_points):
    labels = build_integer_score_labels(max_points)
    matches = list(re.finditer(r"\d+(?:\.\d+)?(?:\s*-\s*\d+(?:\.\d+)?)?\s*分", text))
    if not matches:
        return text
    parts = []
    cursor = 0
    for idx, match in enumerate(matches):
        parts.append(text[cursor:match.start()])
        parts.append(labels[min(idx, len(labels) - 1)])
        cursor = match.end()
    parts.append(text[cursor:])
    return "".join(parts)


def normalize_criteria(criteria, final_points):
    adjusted = []
    for item, new_points in zip(criteria, final_points):
        adjusted.append(
            {
                **item,
                "points": int(new_points),
                "standard": rewrite_standard_scores(item["standard"], int(new_points)),
            }
        )
    return adjusted


def build_dedup_criteria(criteria, configs):
    source_map = {item["subitem"]: item for item in criteria}
    result = []
    for config in configs:
        item = dict(source_map[config["source"]])
        item["points"] = int(config["points"])
        if "category" in config:
            item["category"] = config["category"]
        if "subitem" in config:
            item["subitem"] = config["subitem"]
        if "standard" in config:
            item["standard"] = config["standard"]
        if "evidence" in config:
            item["evidence"] = config["evidence"]
        item["standard"] = rewrite_standard_scores(item["standard"], item["points"])
        result.append(item)
    return result


def build_instruction_sheet(wb):
    ws = wb.create_sheet("说明")
    merged_title(ws, "A1:H1", "软件项目供应商评审评分表（最终版）")

    write_row(
        ws,
        3,
        ["项目", "内容", None, None, None, None, None, None],
        font=font_header,
        fill=fill_header,
        alignment=align_center,
    )
    ws.merge_cells("B3:H3")

    items = [
        ("适用场景", "适用于软件定制开发、研发服务、驻场运维、系统集成配套开发等供应商比选或询比场景。"),
        ("编制原则", "以资格性审查先行、定制开发/驻场运维/统一报价三张评分页分别独立评分、汇总页统一加权为原则编制。"),
        ("资格规则", "资格审查任一强制性条款判定“不通过”，该供应商不进入详细评分和排名。"),
        ("权重规则", "同一项目必须对全部供应商采用统一权重，不得因供应商不同而分别设置权重。"),
        ("价格规则", "默认采用常见的最低有效报价比例法：最低有效报价/本供应商有效报价×价格分值。"),
        ("使用建议", "定制开发评分、驻场运维评分、统一报价评分三页各自满分 100 分；在“汇总排名”页设置三项权重后自动汇总为综合评审得分。"),
    ]
    r = 4
    for key, value in items:
        ws.cell(r, 1, key)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(r, 2, value)
        style_cell(ws.cell(r, 1), font=font_subtitle, fill=fill_section, alignment=align_center)
        style_cell(ws.cell(r, 2), font=font_body, alignment=align_left)
        for c in range(3, 9):
            style_cell(ws.cell(r, c))
        r += 1

    write_row(
        ws,
        11,
        ["使用步骤", None, None, None, None, None, None, None],
        font=font_header,
        fill=fill_header,
        alignment=align_center,
    )
    ws.merge_cells("A11:H11")

    steps = [
        "1. 在“资格审查”页填报各供应商资格结论，蓝色单元格可编辑。",
        "2. 在“定制开发评分”和“驻场运维评分”页分别录入原始分；每页细项分值合计均为 100 分。",
        "3. 在“统一报价评分”页录入各供应商针对本项目整包的有效报价；本页满分 100 分。",
        "4. 在“汇总排名”页统一设置定制开发、驻场运维、统一报价三项权重，三项权重合计应为 100%。",
        "5. 如采购文件另有专门价格评审办法、专门资格条件或评分细则，应在发文前按项目文件统一调整，不得在评审中临时改动。",
    ]
    r = 12
    for step in steps:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(r, 1, step)
        style_cell(ws.cell(r, 1), font=font_body, alignment=align_left)
        r += 1

    write_row(
        ws,
        19,
        ["推荐权重示例", "定制开发权重", "驻场运维权重", "统一报价权重", "权重合计", "适用说明", None, None],
        font=font_header,
        fill=fill_header,
        alignment=align_center,
    )
    example_rows = [
        ("纯定制开发项目", 60, 0, 40, 100, "适用于以需求调研、设计开发、测试上线为主的项目。"),
        ("纯驻场运维项目", 0, 60, 40, 100, "适用于以驻场支持、故障处理、持续服务为主的项目。"),
        ("综合服务项目", 40, 30, 30, 100, "适用于开发建设与驻场保障同时存在的项目。"),
    ]
    rr = 20
    for row in example_rows:
        write_row(ws, rr, [row[0], row[1], row[2], row[3], row[4], row[5], None, None], alignment=align_left)
        for col in range(2, 5):
            ws.cell(rr, col).number_format = "0"
        style_cell(ws.cell(rr, 5), font=font_formula, fill=fill_formula, alignment=align_center, number_format="0")
        ws.merge_cells(start_row=rr, start_column=6, end_row=rr, end_column=8)
        style_cell(ws.cell(rr, 6), font=font_body, alignment=align_left)
        style_cell(ws.cell(rr, 7))
        style_cell(ws.cell(rr, 8))
        rr += 1

    write_row(
        ws,
        25,
        ["建议取证材料", None, None, None, None, None, None, None],
        font=font_header,
        fill=fill_header,
        alignment=align_center,
    )
    ws.merge_cells("A25:H25")
    proofs = [
        "营业执照、税务开票能力证明、信用承诺或信用查询结果。",
        "同类项目合同关键页、验收单、发票或回款证明等业绩支撑材料。",
        "项目经理及核心成员简历、社保或劳动关系证明、岗位分工表。",
        "总体技术方案、实施计划、WBS/里程碑、风险控制方案、服务承诺书。",
        "体系认证、软件著作权、行业资质、保密/安全承诺等证明材料。",
    ]
    rr = 26
    for proof in proofs:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
        ws.cell(rr, 1, f"- {proof}")
        style_cell(ws.cell(rr, 1), font=font_body, alignment=align_left)
        rr += 1

    set_widths(ws, {"A": 18, "B": 14, "C": 14, "D": 14, "E": 12, "F": 20, "G": 14, "H": 14})
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def build_qualification_sheet(wb):
    ws = wb.create_sheet("资格审查")
    merged_title(ws, "A1:F1", "资格性及符合性审查表")
    ws.merge_cells("A2:F2")
    ws["A2"] = "说明：任一强制性条款判定“不通过”，该供应商不得进入详细评分；“待确认”仅用于评审前补证核验，不得用于已终结评审。"
    style_cell(ws["A2"], font=font_body, fill=fill_note, alignment=align_left)

    headers = ["序号", "资格审查项", "审查要求/判定标准", "供应商A", "供应商B", "供应商C"]
    write_row(ws, 4, headers, font=font_header, fill=fill_header, alignment=align_center)

    items = [
        ("主体资格", "依法设立并有效存续，提供营业执照或同等主体资格证明。"),
        ("开票能力", "能够开具合法合规发票，并满足采购文件约定的税务要求。"),
        ("信誉要求", "近三年无重大违法失信记录、无重大质量安全事故、无严重违约情形。"),
        ("技术与服务能力", "具备承担本项目所需的技术、工具、方法、管理和服务能力。"),
        ("最低团队要求", "满足采购文件规定的最低人员配置、驻场安排或响应时效要求。"),
        ("业绩及授权要求", "按采购文件要求提供同类业绩；如涉及代理、原厂授权或联合体要求，应提供有效授权/承诺。"),
        ("商务条款响应", "接受主要合同条款，包括交付、验收、付款、保密、知识产权、违约责任等。"),
        ("安全与保密", "满足网络安全、数据安全、保密及现场管理等强制性要求。"),
        ("材料真实性", "提交的资质、业绩、简历、承诺等材料真实、完整、可核验。"),
        ("其他专门条件", "满足采购文件另行约定的专门资格条件或实质性响应要求。"),
    ]

    start_row = 5
    for idx, (item, rule) in enumerate(items, start=1):
        ws.cell(start_row, 1, idx)
        ws.cell(start_row, 2, item)
        ws.cell(start_row, 3, rule)
        for col in range(1, 7):
            cell = ws.cell(start_row, col)
            style_cell(cell, font=font_body, alignment=align_left if col in (2, 3) else align_center)
        start_row += 1

    dv = DataValidation(type="list", formula1='"通过,不通过,待确认,不适用"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("D5:F14")
    for row in range(5, 15):
        for col in range(4, 7):
            style_cell(ws.cell(row, col), font=font_input, fill=fill_note, alignment=align_center)

    ws.cell(15, 1, "审查结论")
    ws.merge_cells("B15:C15")
    ws["B15"] = "任一项不通过则结论为不通过；存在待确认或空白则结论为待确认；其余视为通过。"
    style_cell(ws["A15"], font=font_subtitle, fill=fill_section, alignment=align_center)
    style_cell(ws["B15"], font=font_body, fill=fill_section, alignment=align_left)
    style_cell(ws["C15"], fill=fill_section)
    for col in range(4, 7):
        col_letter = get_column_letter(col)
        formula = (
            f'=IF(COUNTIF({col_letter}5:{col_letter}14,"不通过")>0,"不通过",'
            f'IF(COUNTIF({col_letter}5:{col_letter}14,"待确认")+COUNTBLANK({col_letter}5:{col_letter}14)>0,"待确认","通过"))'
        )
        ws.cell(15, col, formula)
        style_cell(ws.cell(15, col), font=font_formula, fill=fill_formula, alignment=align_center)

    set_widths(ws, {"A": 8, "B": 18, "C": 54, "D": 14, "E": 14, "F": 14})
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return 15


def build_technical_sheet(wb, title, note, criteria):
    ws = wb.create_sheet(title)
    merged_title(ws, "A1:I1", title)

    ws.merge_cells("A2:I2")
    ws["A2"] = note
    style_cell(ws["A2"], font=font_body, fill=fill_note, alignment=align_left)

    ws.merge_cells("A3:I3")
    ws["A3"] = "填写规则：仅录入蓝色单元格；专家原始分不得超过 C 列分值；本页仅统计技术商务分，不含统一报价得分。"
    style_cell(ws["A3"], font=font_body, alignment=align_left)

    headers = [
        "评审大项",
        "评审子项",
        "分值",
        "评分标准",
        "评分依据/证明材料",
        "供应商A原始分",
        "供应商B原始分",
        "供应商C原始分",
        "备注",
    ]
    write_row(ws, 4, headers, font=font_header, fill=fill_header, alignment=align_center)

    data_start = 5
    current_row = data_start
    for item in criteria:
        values = [item["category"], item["subitem"], item["points"], item["standard"], item["evidence"], None, None, None, None]
        write_row(ws, current_row, values, font=font_body, alignment=align_left)
        style_cell(ws.cell(current_row, 1), fill=fill_section, alignment=align_center)
        style_cell(ws.cell(current_row, 3), alignment=align_center)
        for col in range(6, 9):
            style_cell(ws.cell(current_row, col), font=font_input, fill=fill_note, alignment=align_center, number_format="0.00")
        dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=item["points"], allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"F{current_row}:H{current_row}")
        current_row += 1

    total_points = sum(item["points"] for item in criteria)
    total_row = current_row
    write_row(
        ws,
        total_row,
        ["技术商务小计", None, total_points, "本行自动汇总本页技术商务得分，不含统一报价得分。", None, None, None, None, None],
        font=font_subtitle,
        fill=fill_summary,
        alignment=align_left,
    )
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    for col in range(6, 9):
        cell = ws.cell(total_row, col)
        letter = get_column_letter(col)
        cell.value = f"=ROUND(SUM({letter}{data_start}:{letter}{current_row - 1}),2)"
        style_cell(cell, font=font_formula, fill=fill_formula, alignment=align_center, number_format="0.00")
    style_cell(ws.cell(total_row, 3), font=font_subtitle, fill=fill_summary, alignment=align_center)

    set_widths(
        ws,
        {
            "A": 18,
            "B": 22,
            "C": 8,
            "D": 46,
            "E": 28,
            "F": 13,
            "G": 13,
            "H": 13,
            "I": 18,
        },
    )
    ws.freeze_panes = "F5"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "$1:$4"
    return total_row


def build_price_sheet(wb, price_weight):
    ws = wb.create_sheet(PRICE_SHEET_TITLE)
    merged_title(ws, "A1:I1", PRICE_SHEET_TITLE)

    ws.merge_cells("A2:I2")
    ws["A2"] = f"适用于本项目整包统一报价评分。供应商仅填写一份针对本次项目完整范围的有效报价；本页默认满分 {price_weight} 分。"
    style_cell(ws["A2"], font=font_body, fill=fill_note, alignment=align_left)

    for idx, header in enumerate(["供应商A有效报价", "供应商B有效报价", "供应商C有效报价"], start=10):
        cell = ws.cell(2, idx, header)
        style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)

    ws.merge_cells("A3:I3")
    ws["A3"] = "填写规则：仅录入蓝色单元格；报价应为本项目整包统一有效报价，不再区分定制开发和驻场运维分别报价。"
    style_cell(ws["A3"], font=font_body, alignment=align_left)
    for col in range(10, 13):
        cell = ws.cell(3, col)
        cell.number_format = "0.00"
        style_cell(cell, font=font_input, fill=fill_note, alignment=align_center)
    ws["J3"].comment = Comment("请输入供应商A针对本项目整包的有效报价。", "Codex")
    ws["K3"].comment = Comment("请输入供应商B针对本项目整包的有效报价。", "Codex")
    ws["L3"].comment = Comment("请输入供应商C针对本项目整包的有效报价。", "Codex")

    headers = [
        "评审大项",
        "评审子项",
        "分值",
        "评分标准",
        "评分依据/证明材料",
        "供应商A得分",
        "供应商B得分",
        "供应商C得分",
        "备注",
    ]
    write_row(ws, 4, headers, font=font_header, fill=fill_header, alignment=align_center)
    for col in range(10, 13):
        style_cell(ws.cell(4, col), border_on=False)

    write_row(
        ws,
        5,
        [
            "价格评分",
            "统一报价得分",
            price_weight,
            f"价格部分采用最低有效报价比例法，最低有效报价得满分 {price_weight} 分，其他供应商按“最低有效报价/本供应商有效报价×{price_weight}”折算。",
            "有效报价表、澄清函、报价一览表等。",
            None,
            None,
            None,
            None,
        ],
        font=font_body,
        alignment=align_left,
    )
    style_cell(ws.cell(5, 1), fill=fill_section, alignment=align_center)
    style_cell(ws.cell(5, 3), alignment=align_center)
    for score_col, quote_col in zip(("F", "G", "H"), ("J", "K", "L")):
        formula = (
            f'=IF(OR(${quote_col}$3="",COUNT($J$3:$L$3)=0),0,'
            f'ROUND(MIN($J$3:$L$3)/${quote_col}$3*{price_weight},2))'
        )
        ws[f"{score_col}5"] = formula
        style_cell(ws[f"{score_col}5"], font=font_formula, fill=fill_formula, alignment=align_center, number_format="0.00")
    for col in range(10, 13):
        style_cell(ws.cell(5, col), border_on=False)

    write_row(
        ws,
        6,
        ["合计", None, price_weight, f"本页满分 {price_weight} 分。", None, None, None, None, None],
        font=font_subtitle,
        fill=fill_summary,
        alignment=align_left,
    )
    ws.merge_cells("A6:B6")
    for col in range(6, 9):
        letter = get_column_letter(col)
        ws.cell(6, col, f"=ROUND({letter}5,2)")
        style_cell(ws.cell(6, col), font=font_formula, fill=fill_formula, alignment=align_center, number_format="0.00")
    style_cell(ws.cell(6, 3), font=font_subtitle, fill=fill_summary, alignment=align_center)
    for col in range(10, 13):
        style_cell(ws.cell(6, col), border_on=False)

    set_widths(
        ws,
        {
            "A": 18,
            "B": 22,
            "C": 8,
            "D": 46,
            "E": 28,
            "F": 13,
            "G": 13,
            "H": 13,
            "I": 18,
            "J": 13,
            "K": 13,
            "L": 13,
        },
    )
    ws.freeze_panes = "F5"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "$1:$4"
    return 6


def build_summary_sheet(wb, qual_row, dev_total_row, ops_total_row, price_total_row):
    ws = wb.create_sheet("汇总排名")
    merged_title(ws, "A1:E1", "综合评审汇总及排名")

    write_row(
        ws,
        3,
        ["项目统一权重设置", "数值", None, None, "说明"],
        font=font_header,
        fill=fill_header,
        alignment=align_center,
    )
    ws.merge_cells("B3:D3")

    ws["A4"] = "定制开发权重（%）"
    ws["B4"] = CUSTOM_WEIGHT_DEFAULT
    ws["A5"] = "驻场运维权重（%）"
    ws["B5"] = OPS_WEIGHT_DEFAULT
    ws["A6"] = "统一报价权重（%）"
    ws["B6"] = PRICE_WEIGHT_DEFAULT
    ws["A7"] = "权重合计（%）"
    ws["B7"] = "=SUM(B4:B6)"
    ws["A8"] = "权重使用规则"
    ws["B8"] = "定制开发权重+驻场运维权重+统一报价权重=100%。纯定制开发项目建议 60/0/40，纯驻场项目建议 0/60/40，综合服务项目可按采购要求自行调整。"
    ws.merge_cells("B8:E8")
    for cell_ref in ("A4", "A5", "A6", "A7", "A8"):
        style_cell(ws[cell_ref], font=font_subtitle, fill=fill_section, alignment=align_center)
    style_cell(ws["B4"], font=font_input, fill=fill_note, alignment=align_center, number_format="0")
    style_cell(ws["B5"], font=font_input, fill=fill_note, alignment=align_center, number_format="0")
    style_cell(ws["B6"], font=font_input, fill=fill_note, alignment=align_center, number_format="0")
    style_cell(ws["B7"], font=font_formula, fill=fill_formula, alignment=align_center, number_format="0")
    style_cell(ws["B8"], font=font_body, alignment=align_left)

    weight_dv = DataValidation(type="whole", operator="between", formula1=0, formula2=100, allow_blank=False)
    ws.add_data_validation(weight_dv)
    weight_dv.add("B4")
    weight_dv.add("B5")
    weight_dv.add("B6")

    write_row(
        ws,
        10,
        ["指标", "供应商A", "供应商B", "供应商C", "说明"],
        font=font_header,
        fill=fill_header,
        alignment=align_center,
    )

    rows = {
        11: ("资格审查结论", f"=资格审查!D{qual_row}", f"=资格审查!E{qual_row}", f"=资格审查!F{qual_row}", "仅资格审查“通过”的供应商参与排序。"),
        12: ("定制开发总分", f'=IF(B11="通过",定制开发评分!F{dev_total_row},"")', f'=IF(C11="通过",定制开发评分!G{dev_total_row},"")', f'=IF(D11="通过",定制开发评分!H{dev_total_row},"")', "引用“定制开发评分”页技术商务小计，满分 100 分。"),
        13: ("驻场运维总分", f'=IF(B11="通过",驻场运维评分!F{ops_total_row},"")', f'=IF(C11="通过",驻场运维评分!G{ops_total_row},"")', f'=IF(D11="通过",驻场运维评分!H{ops_total_row},"")', "引用“驻场运维评分”页技术商务小计，满分 100 分。"),
        14: ("统一报价总分", f'=IF(B11="通过",{PRICE_SHEET_TITLE}!F{price_total_row},"")', f'=IF(C11="通过",{PRICE_SHEET_TITLE}!G{price_total_row},"")', f'=IF(D11="通过",{PRICE_SHEET_TITLE}!H{price_total_row},"")', "引用“统一报价评分”页合计分，满分 100 分。"),
        15: ("综合评审得分", '=IF(B11="通过",ROUND(B12*$B$4/100+B13*$B$5/100+B14*$B$6/100,2),"")', '=IF(C11="通过",ROUND(C12*$B$4/100+C13*$B$5/100+C14*$B$6/100,2),"")', '=IF(D11="通过",ROUND(D12*$B$4/100+D13*$B$5/100+D14*$B$6/100,2),"")', "综合得分=定制开发总分×定制开发权重+驻场运维总分×驻场运维权重+统一报价总分×统一报价权重。"),
        16: ("综合排名", '=IF(B11="通过",RANK(B15,$B$15:$D$15,0),"")', '=IF(C11="通过",RANK(C15,$B$15:$D$15,0),"")', '=IF(D11="通过",RANK(D15,$B$15:$D$15,0),"")', "按综合得分由高到低排名，分值相同则并列。"),
    }
    for row, values in rows.items():
        write_row(ws, row, list(values), alignment=align_left)
        style_cell(ws.cell(row, 1), font=font_subtitle if row >= 15 else font_body, fill=fill_section if row >= 15 else None, alignment=align_left if row == 15 else align_center)
        for col in range(2, 5):
            style_cell(
                ws.cell(row, col),
                font=font_formula,
                fill=fill_formula if row >= 12 else None,
                alignment=align_center,
                number_format="0.00" if row in (12, 13, 14, 15) else None,
            )

    write_row(
        ws,
        19,
        ["补充说明", None, None, None, None],
        font=font_header,
        fill=fill_header,
        alignment=align_center,
    )
    ws.merge_cells("A19:E19")
    notes = [
        "1. 如项目只适用一种评分页，可将其他评分页权重设置为 0。",
        "2. 请确保 B4:B6 三项权重合计为 100；B7 为自动校验值。",
        "3. 如采购文件采用基准价法或其他价格评审办法，应在“统一报价评分”页统一替换价格公式。",
        "4. 本模板为通用正式版，项目特有的资质门槛、涉密要求、原厂授权要求应根据采购文件另行补充。",
    ]
    row = 20
    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row, 1, note)
        style_cell(ws.cell(row, 1), font=font_body, alignment=align_left)
        row += 1

    set_widths(ws, {"A": 20, "B": 14, "C": 14, "D": 14, "E": 48})
    ws.freeze_panes = "A10"
    ws.sheet_view.showGridLines = False


def build_workbook():
    dev_criteria = [
        {"category": "项目理解与需求分析", "subitem": "需求理解准确性", "points": 4, "standard": "对项目背景、建设目标、核心需求理解准确，问题识别到位，得 4 分；理解较为准确但分析深度一般，得 2-3 分；理解偏差明显或内容空泛，得 0-1 分。", "evidence": "需求响应、现状分析、问题清单、项目目标说明。"},
        {"category": "项目理解与需求分析", "subitem": "业务场景与目标匹配度", "points": 3, "standard": "能够结合采购场景提出针对性理解和目标拆解，得 3 分；匹配度一般，得 1-2 分；针对性不足，得 0 分。", "evidence": "业务场景分析、用户画像、流程梳理成果。"},
        {"category": "项目理解与需求分析", "subitem": "范围边界与关键难点识别", "points": 3, "standard": "对项目边界、接口边界和关键难点识别完整，得 3 分；识别基本到位，得 1-2 分；识别不足，得 0 分。", "evidence": "边界清单、关键风险点、约束条件说明。"},
        {"category": "技术方案与架构设计", "subitem": "总体架构与技术路线合理性", "points": 5, "standard": "总体架构完整、技术路线成熟可行、与项目目标高度匹配，得 5 分；较为合理，得 3-4 分；合理性较弱，得 0-2 分。", "evidence": "总体架构图、技术选型说明、部署方案。"},
        {"category": "技术方案与架构设计", "subitem": "功能方案完整性", "points": 4, "standard": "功能设计覆盖主要需求并体现关键场景，得 4 分；覆盖较完整，得 2-3 分；缺项明显，得 0-1 分。", "evidence": "功能清单、原型图、模块设计说明。"},
        {"category": "技术方案与架构设计", "subitem": "接口、数据与安全设计", "points": 3, "standard": "接口衔接、数据流向、安全控制设计清楚，得 3 分；基本具备，得 1-2 分；说明不足，得 0 分。", "evidence": "接口清单、数据模型、安全方案。"},
        {"category": "技术方案与架构设计", "subitem": "可扩展性与可实施性", "points": 3, "standard": "方案具备扩展性、兼容性和可实施路径，得 3 分；一般，得 1-2 分；可实施性弱，得 0 分。", "evidence": "实施路径、环境适配说明、扩展设计。"},
        {"category": "实施计划与项目管理", "subitem": "实施计划与里程碑设置", "points": 4, "standard": "阶段划分、里程碑、关键交付节点清晰可执行，得 4 分；较清晰，得 2-3 分；安排粗略，得 0-1 分。", "evidence": "WBS、里程碑计划、甘特图。"},
        {"category": "实施计划与项目管理", "subitem": "风险控制与质量保证", "points": 3, "standard": "具备完整的风险识别、预警、应对和质量控制机制，得 3 分；机制基本可行，得 1-2 分；机制缺失，得 0 分。", "evidence": "风险台账、测试策略、质量控制流程。"},
        {"category": "实施计划与项目管理", "subitem": "交付物定义与验收安排", "points": 3, "standard": "交付物清单、输出形式、验收方式明确，得 3 分；基本明确，得 1-2 分；不明确，得 0 分。", "evidence": "交付物清单、验收计划、文档目录。"},
        {"category": "团队配置与履约能力", "subitem": "项目经理与核心骨干匹配度", "points": 4, "standard": "项目经理及核心成员经验丰富、履历与项目高度匹配，得 4 分；基本匹配，得 2-3 分；匹配度较低，得 0-1 分。", "evidence": "人员简历、项目经历、证书、社保证明。"},
        {"category": "团队配置与履约能力", "subitem": "岗位配置完整性", "points": 3, "standard": "产品、设计、开发、测试、实施等关键岗位配置齐全且职责清晰，得 3 分；基本齐备，得 1-2 分；明显缺失，得 0 分。", "evidence": "组织架构图、岗位分工、人员清单。"},
        {"category": "团队配置与履约能力", "subitem": "驻场/协同与资源保障能力", "points": 3, "standard": "驻场计划、协同机制、替补资源保障合理，得 3 分；一般，得 1-2 分；保障不足，得 0 分。", "evidence": "驻场计划、协同机制、替补安排。"},
        {"category": "同类业绩与资质能力", "subitem": "同类案例数量与相关性", "points": 5, "standard": "同类案例数量充足、与本项目行业和复杂度高度相关，得 5 分；相关性一般，得 2-4 分；较弱，得 0-1 分。", "evidence": "合同关键页、验收证明、发票或回款证明。"},
        {"category": "同类业绩与资质能力", "subitem": "业绩证明完整性", "points": 2, "standard": "案例证明材料齐全、可核验，得 2 分；部分缺失，得 1 分；证明不足，得 0 分。", "evidence": "合同、验收、佐证清单。"},
        {"category": "同类业绩与资质能力", "subitem": "专业资质与质量体系", "points": 3, "standard": "具备与项目匹配的体系认证或专业资质，且有效性明确，得 3 分；资质较少或匹配度一般，得 1-2 分；无有效资质，得 0 分。", "evidence": "ISO、CMMI、软件企业能力或行业资质证明。"},
        {"category": "服务承诺与文件质量", "subitem": "售后服务、响应与质保承诺", "points": 3, "standard": "售后机制、响应时效、质保范围和服务边界明确，得 3 分；一般，得 1-2 分；承诺不明确，得 0 分。", "evidence": "售后服务方案、承诺函、SLA。"},
        {"category": "服务承诺与文件质量", "subitem": "响应文件完整性与规范性", "points": 2, "standard": "文件结构清楚、内容完整、查阅方便、格式规范，得 2 分；一般，得 1 分；较差，得 0 分。", "evidence": "响应文件目录、排版、盖章签字情况。"},
    ]

    ops_criteria = [
        {"category": "服务理解与方案设计", "subitem": "服务范围与边界理解", "points": 4, "standard": "对服务范围、职责边界、支持边界理解准确，得 4 分；基本准确，得 2-3 分；理解偏差明显，得 0-1 分。", "evidence": "服务方案、范围说明、职责分工。"},
        {"category": "服务理解与方案设计", "subitem": "服务流程与报告机制", "points": 3, "standard": "服务流程、工单流转、例会与报告机制完整，得 3 分；基本具备，得 1-2 分；机制不足，得 0 分。", "evidence": "服务流程图、报告模板、例会机制。"},
        {"category": "服务理解与方案设计", "subitem": "服务方案可执行性", "points": 3, "standard": "方案执行路径清晰、适配采购场景，得 3 分；一般，得 1-2 分；可执行性弱，得 0 分。", "evidence": "实施路径、现场协同方案。"},
        {"category": "驻场与响应保障", "subitem": "驻场人员数量与岗位结构", "points": 6, "standard": "驻场人员数量充分、岗位结构合理、关键岗位齐备，得 6 分；基本满足，得 3-5 分；明显不足，得 0-2 分。", "evidence": "驻场清单、岗位说明、资源投入计划。"},
        {"category": "驻场与响应保障", "subitem": "人员稳定性与替补机制", "points": 4, "standard": "稳定性保障和替补机制明确，得 4 分；较为明确，得 2-3 分；安排不足，得 0-1 分。", "evidence": "替补名单、稳定性承诺、人员管理机制。"},
        {"category": "驻场与响应保障", "subitem": "响应时效与故障升级机制", "points": 6, "standard": "响应时效、升级路径、闭环机制明确且优于基本要求，得 6 分；基本满足，得 3-5 分；较弱，得 0-2 分。", "evidence": "SLA、升级流程、值守安排。"},
        {"category": "驻场与响应保障", "subitem": "应急保障与重大问题处置能力", "points": 4, "standard": "应急预案、重大故障处置、专项支撑能力强，得 4 分；一般，得 2-3 分；较弱，得 0-1 分。", "evidence": "应急预案、值班机制、历史案例说明。"},
        {"category": "持续服务与知识交接", "subitem": "知识库、文档与培训交接安排", "points": 5, "standard": "知识沉淀、文档维护、培训交接安排完整，得 5 分；较完整，得 3-4 分；较弱，得 0-2 分。", "evidence": "知识库目录、交接计划、培训方案。"},
        {"category": "持续服务与知识交接", "subitem": "持续优化与例会复盘机制", "points": 5, "standard": "具备持续优化、例会复盘、改进闭环机制，得 5 分；一般，得 3-4 分；机制不足，得 0-2 分。", "evidence": "复盘模板、改进机制、周/月报样例。"},
        {"category": "同类经验与资质能力", "subitem": "同类服务案例数量与相关性", "points": 6, "standard": "同类服务案例丰富且相关性高，得 6 分；一般，得 3-5 分；较弱，得 0-2 分。", "evidence": "服务合同、验收证明、用户评价或成果材料。"},
        {"category": "同类经验与资质能力", "subitem": "资质与服务管理体系", "points": 4, "standard": "具备与服务管理相关的体系认证或有效资质，得 4 分；一般，得 2-3 分；无有效资质，得 0-1 分。", "evidence": "ISO20000、ISO27001、质量体系等。"},
        {"category": "商务响应与文件质量", "subitem": "商务承诺与违约响应机制", "points": 4, "standard": "商务承诺明确、违约处理和资源补偿机制清晰，得 4 分；一般，得 2-3 分；较弱，得 0-1 分。", "evidence": "商务承诺函、违约处理条款响应。"},
        {"category": "商务响应与文件质量", "subitem": "文件完整性与规范性", "points": 6, "standard": "文件完整、结构清晰、排版规范、检索方便，得 6 分；一般，得 3-5 分；较弱，得 0-2 分。", "evidence": "响应文件目录、盖章签字、版本控制。"},
    ]

    existing = load_existing_inputs(len(dev_criteria), len(ops_criteria))
    dev_configs = [
        {"source": "需求理解准确性", "points": 8},
        {"source": "业务场景与目标匹配度", "points": 7},
        {"source": "范围边界与关键难点识别", "points": 6},
        {"source": "总体架构与技术路线合理性", "points": 10},
        {"source": "功能方案完整性", "points": 9},
        {"source": "接口、数据与安全设计", "points": 8},
        {"source": "可扩展性与可实施性", "points": 7},
        {"source": "实施计划与里程碑设置", "points": 8},
        {"source": "风险控制与质量保证", "points": 7},
        {"source": "交付物定义与验收安排", "points": 6},
        {"source": "项目经理与核心骨干匹配度", "points": 8},
        {"source": "岗位配置完整性", "points": 6, "subitem": "开发团队岗位配置完整性", "evidence": "组织架构图、开发/测试/实施岗位分工、人员清单。"},
        {"source": "同类案例数量与相关性", "points": 6, "category": "软件交付经验与技术能力", "subitem": "同类软件交付案例数量与相关性", "evidence": "软件开发或系统建设项目合同关键页、验收证明、成果材料。"},
        {"source": "专业资质与质量体系", "points": 4, "category": "软件交付经验与技术能力", "subitem": "技术资质与研发能力", "evidence": "软件研发类资质、软著、专利、技术能力证明。"},
    ]
    ops_configs = [
        {"source": "服务范围与边界理解", "points": 7},
        {"source": "服务流程与报告机制", "points": 6},
        {"source": "服务方案可执行性", "points": 6},
        {"source": "驻场人员数量与岗位结构", "points": 12},
        {"source": "人员稳定性与替补机制", "points": 8},
        {"source": "响应时效与故障升级机制", "points": 12},
        {"source": "应急保障与重大问题处置能力", "points": 8},
        {"source": "知识库、文档与培训交接安排", "points": 7},
        {"source": "持续优化与例会复盘机制", "points": 7},
        {"source": "同类服务案例数量与相关性", "points": 8, "category": "驻场服务经验与治理能力", "subitem": "同类驻场服务案例数量与相关性", "evidence": "驻场运维或长期服务项目合同、验收证明、用户评价材料。"},
        {"source": "资质与服务管理体系", "points": 7, "category": "驻场服务经验与治理能力", "subitem": "服务管理体系与SLA治理能力", "evidence": "ISO20000、服务管理制度、SLA治理机制说明。"},
        {"source": "商务承诺与违约响应机制", "points": 7},
        {"source": "文件完整性与规范性", "points": 5, "subitem": "响应文件完整性与规范性"},
    ]
    dev_criteria = build_dedup_criteria(dev_criteria, dev_configs)
    ops_criteria = build_dedup_criteria(ops_criteria, ops_configs)

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation = CalcProperties(calcMode="auto")

    build_instruction_sheet(wb)
    qual_row = build_qualification_sheet(wb)

    dev_total_row = build_technical_sheet(
        wb,
        "定制开发评分",
        "适用于以需求调研、系统设计、开发测试、上线交付为主的软件项目。本页按“开发特有能力”为主设置评分项，已对与驻场运维页重复的通用项做去重处理，页内满分 100 分。",
        dev_criteria,
    )
    ops_total_row = build_technical_sheet(
        wb,
        "驻场运维评分",
        "适用于以驻场研发、运维支持、持续服务和响应保障为主的项目。本页按“驻场运维特有能力”为主设置评分项，已对与定制开发页重复的通用项做去重处理，页内满分 100 分。",
        ops_criteria,
    )
    price_total_row = build_price_sheet(wb, 100)

    build_summary_sheet(wb, qual_row, dev_total_row, ops_total_row, price_total_row)
    apply_existing_inputs(wb, existing)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(DEDUP_OUTPUT)
        return DEDUP_OUTPUT
    except PermissionError:
        wb.save(DEDUP_FALLBACK_OUTPUT)
        return DEDUP_FALLBACK_OUTPUT


if __name__ == "__main__":
    path = build_workbook()
    print(path)
