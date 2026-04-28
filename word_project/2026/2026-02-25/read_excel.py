import openpyxl

wb = openpyxl.load_workbook(r'd:\DevelopmentLocation\agent skill\skills\word_project\2026\2026-02-25\经营管理分析.xlsx')
ws = wb.active

hidden_rows = set()
for row_idx in range(1, ws.max_row + 1):
    rd = ws.row_dimensions.get(row_idx)
    if rd and rd.hidden:
        hidden_rows.add(row_idx)

output = []
output.append(f"Hidden rows: {sorted(hidden_rows)}")

for row_idx in range(1, ws.max_row + 1):
    if row_idx in hidden_rows:
        continue
    a_val = ws.cell(row_idx, 1).value
    b_val = ws.cell(row_idx, 2).value
    c_val = ws.cell(row_idx, 3).value
    d_val = ws.cell(row_idx, 4).value
    e_val = ws.cell(row_idx, 5).value

    if any([a_val, b_val, c_val, d_val, e_val]):
        output.append(f"\n=== Row {row_idx} ===")
        if a_val: output.append(f"  A(模块): {str(a_val)}")
        if b_val: output.append(f"  B(功能): {str(b_val)}")
        if c_val: output.append(f"  C(描述): {str(c_val)}")
        if d_val: output.append(f"  D(详细): {str(d_val)}")
        if e_val: output.append(f"  E(备注): {str(e_val)}")

with open(r'd:\DevelopmentLocation\agent skill\skills\word_project\2026\2026-02-25\excel_data.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print("Done! Written to excel_data.txt")
print(f"Total visible rows with data: {len([x for x in output if x.startswith('=== Row')])}")
