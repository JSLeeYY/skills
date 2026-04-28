from copy import copy
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(r"D:\DevelopmentLocation\agent skill\skills\word_project\2026\2026-04-23")
SOURCE = BASE / "软件项目供应商评审评分表-去重版.xlsx"
TARGET = BASE / "软件项目供应商评审评分表-评分项优化版.xlsx"


def cell_style_snapshot(ws, row, cols):
    out = {}
    for c in cols:
        cell = ws.cell(row, c)
        out[c] = {
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
        }
    return out


def apply_style(cell, snapshot):
    cell.font = copy(snapshot["font"])
    cell.fill = copy(snapshot["fill"])
    cell.border = copy(snapshot["border"])
    cell.alignment = copy(snapshot["alignment"])
    cell.number_format = snapshot["number_format"]
    cell.protection = copy(snapshot["protection"])


def clear_data_area(ws, start_row):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= start_row:
            ws.unmerge_cells(str(merged))
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def rebuild_score_sheet(ws, items, group_specs, subtotal_title, subtotal_note, subtotal_row_style_src):
    cols = range(1, 10)
    data_style = cell_style_snapshot(ws, 5, cols)
    subtotal_style = cell_style_snapshot(ws, subtotal_row_style_src, cols)

    row_heights = {}
    for item in items:
        total = 0
        has_height = False
        for r in item["source_rows"]:
            h = ws.row_dimensions[r].height
            if h:
                total += h
                has_height = True
        row_heights[item["name"]] = total if has_height else None

    subtotal_height = ws.row_dimensions[subtotal_row_style_src].height

    clear_data_area(ws, 5)

    current_row = 5
    group_start_rows = {}
    for item in items:
        group_start_rows.setdefault(item["major"], current_row)

        for c in cols:
            apply_style(ws.cell(current_row, c), data_style[c])

        ws.cell(current_row, 2).value = item["name"]
        ws.cell(current_row, 3).value = item["score"]
        ws.cell(current_row, 4).value = item["standard"]
        ws.cell(current_row, 5).value = item["basis"]
        ws.cell(current_row, 6).value = None
        ws.cell(current_row, 7).value = None
        ws.cell(current_row, 8).value = None
        ws.cell(current_row, 9).value = None

        if row_heights[item["name"]]:
            ws.row_dimensions[current_row].height = row_heights[item["name"]]
        current_row += 1

    subtotal_row = current_row
    for c in cols:
        apply_style(ws.cell(subtotal_row, c), subtotal_style[c])
    ws.cell(subtotal_row, 1).value = subtotal_title
    ws.cell(subtotal_row, 3).value = 100
    ws.cell(subtotal_row, 4).value = subtotal_note
    ws.cell(subtotal_row, 6).value = f"=ROUND(SUM(F5:F{subtotal_row - 1}),2)"
    ws.cell(subtotal_row, 7).value = f"=ROUND(SUM(G5:G{subtotal_row - 1}),2)"
    ws.cell(subtotal_row, 8).value = f"=ROUND(SUM(H5:H{subtotal_row - 1}),2)"
    if subtotal_height:
        ws.row_dimensions[subtotal_row].height = subtotal_height

    last_keep_row = subtotal_row
    if ws.max_row > last_keep_row:
        ws.delete_rows(last_keep_row + 1, ws.max_row - last_keep_row)

    for major, count in group_specs:
        start = group_start_rows[major]
        end = start + count - 1
        ws.cell(start, 1).value = major
        if count > 1:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)

    return subtotal_row


wb = load_workbook(SOURCE)

custom_ws = wb["定制开发评分"]
onsite_ws = wb["驻场运维评分"]
summary_ws = wb["汇总排名"]

custom_items = [
    {
        "major": "项目理解与需求分析",
        "name": "需求理解与场景匹配度",
        "score": 15,
        "standard": "对项目背景、建设目标、核心需求理解准确，能够结合采购场景提出针对性分析和目标拆解，得 13-15 分；理解较为准确、匹配度较好，得 8-12 分；理解不准确或针对性不足，得 0-7 分。",
        "basis": "需求响应、现状分析、问题清单、项目目标说明、业务场景分析、用户画像、流程梳理成果。",
        "source_rows": [5, 6],
    },
    {
        "major": "项目理解与需求分析",
        "name": "范围边界与关键难点识别",
        "score": 6,
        "standard": "对项目边界、接口边界和关键难点识别完整，得 6 分；识别基本到位，得 3-5 分；识别不足，得 0-2 分。",
        "basis": "边界清单、关键风险点、约束条件说明。",
        "source_rows": [7],
    },
    {
        "major": "技术方案与架构设计",
        "name": "总体架构与功能方案合理性",
        "score": 19,
        "standard": "总体架构完整、技术路线成熟，功能设计覆盖主要需求并体现关键场景，得 16-19 分；方案较完整合理，得 10-15 分；缺项较多或合理性较弱，得 0-9 分。",
        "basis": "总体架构图、技术选型说明、部署方案、功能清单、原型图、模块设计说明。",
        "source_rows": [8, 9],
    },
    {
        "major": "技术方案与架构设计",
        "name": "接口、数据与安全设计",
        "score": 8,
        "standard": "接口衔接、数据流向、安全控制设计清楚，得 8 分；基本具备，得 4-7 分；说明不足，得 0-3 分。",
        "basis": "接口清单、数据模型、安全方案。",
        "source_rows": [10],
    },
    {
        "major": "技术方案与架构设计",
        "name": "可扩展性与可实施性",
        "score": 7,
        "standard": "方案具备扩展性、兼容性和可实施路径，得 7 分；一般，得 4-6 分；可实施性较弱，得 0-3 分。",
        "basis": "实施路径、环境适配说明、扩展设计。",
        "source_rows": [11],
    },
    {
        "major": "实施计划与项目管理",
        "name": "实施计划、风险与质量控制",
        "score": 15,
        "standard": "实施阶段划分、里程碑设置、风险识别、预警应对和质量控制机制完整，得 13-15 分；安排较清晰，机制基本可行，得 8-12 分；安排粗略或机制不足，得 0-7 分。",
        "basis": "WBS、里程碑计划、甘特图、风险台账、测试策略、质量控制流程。",
        "source_rows": [12, 13],
    },
    {
        "major": "实施计划与项目管理",
        "name": "交付物定义与验收安排",
        "score": 6,
        "standard": "交付物清单、输出形式、验收方式明确，得 6 分；基本明确，得 3-5 分；不明确，得 0-2 分。",
        "basis": "交付物清单、验收计划、文档目录。",
        "source_rows": [14],
    },
    {
        "major": "团队配置与履约能力",
        "name": "项目团队配置与履约能力",
        "score": 14,
        "standard": "项目经理及核心成员经验丰富，团队岗位配置齐全、职责清晰，得 12-14 分；基本匹配，得 7-11 分；团队配置或履约能力较弱，得 0-6 分。",
        "basis": "人员简历、项目经历、证书、社保证明、组织架构图、开发/测试/实施岗位分工、人员清单。",
        "source_rows": [15, 16],
    },
    {
        "major": "软件交付经验与技术能力",
        "name": "同类软件交付经验与技术能力",
        "score": 10,
        "standard": "同类软件交付案例数量充足、相关性高，且具备与项目匹配的技术资质或研发能力，得 9-10 分；相关性和技术能力一般，得 5-8 分；较弱，得 0-4 分。",
        "basis": "软件开发或系统建设项目合同关键页、验收证明、成果材料、软件研发类资质、软著、专利、技术能力证明。",
        "source_rows": [17, 18],
    },
]

custom_groups = [
    ("项目理解与需求分析", 2),
    ("技术方案与架构设计", 3),
    ("实施计划与项目管理", 2),
    ("团队配置与履约能力", 1),
    ("软件交付经验与技术能力", 1),
]

onsite_items = [
    {
        "major": "服务理解与方案设计",
        "name": "服务范围与流程理解",
        "score": 13,
        "standard": "对服务范围、职责边界、支持边界理解准确，服务流程、工单流转、例会与报告机制完整，得 11-13 分；基本具备，得 7-10 分；理解偏差明显或机制不足，得 0-6 分。",
        "basis": "服务方案、范围说明、职责分工、服务流程图、报告模板、例会机制。",
        "source_rows": [5, 6],
    },
    {
        "major": "服务理解与方案设计",
        "name": "服务方案可执行性",
        "score": 6,
        "standard": "方案执行路径清晰、适配采购场景，得 6 分；一般，得 3-5 分；可执行性较弱，得 0-2 分。",
        "basis": "实施路径、现场协同方案。",
        "source_rows": [7],
    },
    {
        "major": "驻场与响应保障",
        "name": "驻场人员配置与稳定性保障",
        "score": 20,
        "standard": "驻场人员数量充分、岗位结构合理、关键岗位齐备，且人员稳定性保障、替补机制明确，得 17-20 分；基本满足，得 10-16 分；明显不足，得 0-9 分。",
        "basis": "驻场清单、岗位说明、资源投入计划、替补名单、稳定性承诺、人员管理机制。",
        "source_rows": [8, 9],
    },
    {
        "major": "驻场与响应保障",
        "name": "响应时效与应急处置能力",
        "score": 20,
        "standard": "响应时效、升级路径、闭环机制明确，应急预案、重大故障处置和专项支撑能力强，得 17-20 分；基本满足，得 10-16 分；较弱，得 0-9 分。",
        "basis": "SLA、升级流程、值守安排、应急预案、值班机制、历史案例说明。",
        "source_rows": [10, 11],
    },
    {
        "major": "持续服务与知识交接",
        "name": "知识交接与持续优化机制",
        "score": 14,
        "standard": "知识沉淀、文档维护、培训交接安排完整，且具备持续优化、例会复盘、改进闭环机制，得 12-14 分；较完整，得 7-11 分；较弱，得 0-6 分。",
        "basis": "知识库目录、交接计划、培训方案、复盘模板、改进机制、周/月报样例。",
        "source_rows": [12, 13],
    },
    {
        "major": "驻场服务经验与治理能力",
        "name": "同类驻场服务案例与治理能力",
        "score": 15,
        "standard": "同类服务案例丰富、相关性高，且具备与服务管理相关的体系认证或有效资质，得 13-15 分；一般，得 8-12 分；较弱，得 0-7 分。",
        "basis": "驻场运维或长期服务项目合同、验收证明、用户评价材料、ISO20000、服务管理制度、SLA 治理机制说明。",
        "source_rows": [14, 15],
    },
    {
        "major": "商务响应与文件质量",
        "name": "商务响应与文件完整性",
        "score": 12,
        "standard": "商务承诺明确、违约处理和资源补偿机制清晰，且文件完整、结构清晰、排版规范，得 10-12 分；一般，得 6-9 分；较弱，得 0-5 分。",
        "basis": "商务承诺函、违约处理条款响应、响应文件目录、盖章签字、版本控制。",
        "source_rows": [16, 17],
    },
]

onsite_groups = [
    ("服务理解与方案设计", 2),
    ("驻场与响应保障", 2),
    ("持续服务与知识交接", 1),
    ("驻场服务经验与治理能力", 1),
    ("商务响应与文件质量", 1),
]

custom_total_row = rebuild_score_sheet(
    custom_ws,
    custom_items,
    custom_groups,
    "技术商务小计",
    "本行自动汇总本页技术商务得分，不含统一报价得分。",
    19,
)

onsite_total_row = rebuild_score_sheet(
    onsite_ws,
    onsite_items,
    onsite_groups,
    "技术商务小计",
    "本行自动汇总本页技术商务得分，不含统一报价得分。",
    18,
)

# Update summary references
summary_ws["B11"] = f'=IF(B10="通过",定制开发评分!F{custom_total_row},"")'
summary_ws["C11"] = f'=IF(C10="通过",定制开发评分!G{custom_total_row},"")'
summary_ws["D11"] = f'=IF(D10="通过",定制开发评分!H{custom_total_row},"")'

summary_ws["B12"] = f'=IF(B10="通过",驻场运维评分!F{onsite_total_row},"")'
summary_ws["C12"] = f'=IF(C10="通过",驻场运维评分!G{onsite_total_row},"")'
summary_ws["D12"] = f'=IF(D10="通过",驻场运维评分!H{onsite_total_row},"")'

wb.save(TARGET)
print(TARGET)
